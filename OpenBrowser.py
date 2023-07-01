from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from DownloadCaptcha import audio
from ReadCaptcha import read
import time
import os
import openpyxl

wb = openpyxl.load_workbook("D:\\Downloads\\Book1.xlsx")
sh1 = wb['Sheet1']
row = sh1.max_row

for k in range(2, row + 1):
    GSTIN = sh1.cell(k, 1).value
    s = Service("D:\\Downloads\\chromedriver_win32\\chromedriver.exe")
    driver = webdriver.Chrome(service=s)

    driver.get("https://services.gst.gov.in/services/searchtp")
    driver.implicitly_wait(2)
    driver.find_element(By.ID, 'for_gstin').send_keys(GSTIN)
    time.sleep(2)
    button = driver.find_element(By.XPATH,
                                 "/html/body/div[2]/div[2]/div/div[2]/div/div/form/div[3]/div/div/div/table/tbody/tr[1]/th[2]/button")
    driver.implicitly_wait(2)
    button.click()

    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[1])
    driver.get("https://services.gst.gov.in//services//audiocaptcha")
    act = ActionChains(driver)
    time.sleep(2)
    location = audio(act)
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    captcha = read()
    driver.implicitly_wait(2)
    driver.find_element(By.ID, 'fo-captcha').send_keys(captcha)
    time.sleep(2)
    driver.find_element(By.ID, 'lotsearch').click()
    os.remove('C://Users/Piyush/Downloads/audiocaptcha.mp3')

    for element in driver.find_elements(By.XPATH, '//*[@id="lottable"]/div[2]/div[1]/div/div[1]/p[2]'):
        name = element.text
    try:
        result = name
        if result != sh1.cell(k-1,2).value:
            print(result,k)
            sh1.cell(k, 2, value=result)
            wb.save('D:\\Downloads\\Book1.xlsx')
            result = "Invalid Captcha"
            driver.close()
        else:
            try:
                for element in driver.find_elements(By.XPATH,
                                                    '/html/body/div[2]/div[2]/div/div[2]/div/div/form/div[2]/div/div/span'):
                    name = element.text
                if len(name) > 0:
                    print('Invalid Captcha', k)
                    sh1.cell(k, 2, value='Invalid Captcha')
                    wb.save('D:\\Downloads\\Book1.xlsx')
                    driver.close()
            except:
                for element in driver.find_elements(By.XPATH,
                                                    '/html/body/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div/span[1]'):
                    name = element.text
                if len(name) > 0:
                    print('Invalid GSTIN', k)
                    sh1.cell(k, 2, value='Invalid GSTIN')
                    wb.save('D:\\Downloads\\Book1.xlsx')
                    driver.close()
    except:
        try:
            for element in driver.find_elements(By.XPATH,
                                                '/html/body/div[2]/div[2]/div/div[2]/div/div/form/div[2]/div/div/span'):
                name = element.text
            if len(name) > 0:
                print('Invalid Captcha', k)
                sh1.cell(k, 2, value='Invalid Captcha')
                wb.save('D:\\Downloads\\Book1.xlsx')
                driver.close()
        except:
            print('Invalid GSTIN', k)
            sh1.cell(k, 2, value='Invalid GSTIN')
            wb.save('D:\\Downloads\\Book1.xlsx')
            driver.close()
